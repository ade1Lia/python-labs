
from openpyxl import load_workbook

file_path = 'table.xlsx'
wb = load_workbook(file_path)
ws = wb.active

average_salary = sum([
    ws.cell(row=4, column=6).value,
    ws.cell(row=9, column=6).value,
    ws.cell(row=11, column=6).value
]) / 3

salaries = [ws.cell(row=i, column=6).value for i in range(2, 10) if i not in [4, 9]]
minimum_salary = min(salaries)
maximum_salary = max(salaries)

ws["B14"] = "Средняя зарплата по отделам:"
ws["C14"] = average_salary
ws["B15"] = "Максимальная зарплата:"
ws["C15"] = maximum_salary
ws["B16"] = "Минимальная зарплата:"
ws["C16"] = minimum_salary

wb.save("table.xlsx")
