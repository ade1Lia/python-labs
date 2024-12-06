
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference

file_path = 'table.xlsx'
wb = load_workbook(file_path)
ws = wb.active

departments = {}

department_rows = [4, 9, 11]
for row in department_rows:
    department_name = ws.cell(row=row, column=3).value.split('\n')[0]
    total_salary = ws.cell(row=row, column=6).value
    departments[department_name] = total_salary

ws.append(["Отдел", "Сумма зарплаты"])
for dept, salary in departments.items():
    ws.append([dept, salary])

chart = PieChart()
chart.title = "Распределение зарплат по отделам"

labels = Reference(ws, min_col=1, min_row=ws.max_row - len(departments), max_row=ws.max_row)
data = Reference(ws, min_col=2, min_row=ws.max_row - len(departments), max_row=ws.max_row)

chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

ws.add_chart(chart, "K1")

output_file_path = 'table.xlsx'
wb.save(output_file_path)
